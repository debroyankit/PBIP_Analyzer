"""Export a DependencyGraph to a formatted Excel workbook (.xlsx).

Generates exactly three sheets:

    Sheet 1 — Visual Inventory
        One row per visual. Shows direct measures, dependency tables/columns
        (including transitive lineage through measure chains and calculated
        columns), and the full DAX of every direct measure.

    Sheet 2 — Impact Analysis
        One row per Table, Measure, and Column. Provides objective usage counts,
        relationship participation, and a factual status classification to help
        analysts understand the impact surface of each entity.

    Sheet 3 — Measure Lineage
        One row per measure showing transitive base columns, dependency tables,
        visual usage, measure-to-measure lineage, and the full DAX expression.

The workbook is always saved to the user's Downloads folder automatically.

Requires: openpyxl  (pip install openpyxl)
"""

from __future__ import annotations

import io
import os
import re
from datetime import datetime
from pathlib import Path
from typing import Union

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from services.dependency_engine import (
    DependencyGraph,
    find_unused_entities,
    trace_measure_lineage,
)
from utils.logging_config import get_logger

logger = get_logger("excel_export")

# ---------------------------------------------------------------------------
# Colour palette  (deep navy header, white text, clean body)
# ---------------------------------------------------------------------------
_HEADER_FILL  = PatternFill("solid", fgColor="1F3864")   # deep navy
_HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_BODY_FONT    = Font(name="Calibri", size=10)
_WRAP_ALIGN   = Alignment(wrap_text=True, vertical="top")
_TOP_ALIGN    = Alignment(vertical="top")
_CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Subtle highlight for "No References Found" rows (pale warm yellow)
_NO_REF_FILL  = PatternFill("solid", fgColor="FFF8E1")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _is_system(name: str) -> bool:
    return name.startswith("LocalDateTable_") or name.startswith("DateTableTemplate_")


def _join(items: set | list, sep: str = ", ") -> str:
    """Convert a set/list to a sorted, deduplicated comma-separated string.
    Returns empty string (not 'None' or '[]') when there are no items."""
    cleaned = sorted({str(i) for i in items if i})
    return sep.join(cleaned) if cleaned else ""


def _val(v: str) -> str | None:
    """Return None for empty strings so cells stay blank in Excel."""
    return v if v else None


def _style_header(ws, n_cols: int) -> None:
    """Apply navy header style, freeze the top row, and enable auto-filter."""
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = _CENTER_ALIGN
    ws.freeze_panes   = "A2"
    ws.auto_filter.ref = ws.dimensions


def _auto_width(ws, min_w: int = 10, max_w: int = 65) -> None:
    """Auto-fit column widths based on the longest line in each column."""
    for col_cells in ws.columns:
        best = min_w
        for cell in col_cells:
            if cell.value is not None:
                longest = max(str(cell.value).split("\n"), key=len, default="")
                best = max(best, min(len(longest) + 4, max_w))
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = best


def _write_row(ws, row_idx: int, values: list, wrap_cols: set[int] | None = None) -> None:
    for col_idx, val in enumerate(values, 1):
        cell       = ws.cell(row=row_idx, column=col_idx, value=_val(str(val)) if val is not None else None)
        cell.font  = _BODY_FONT
        if wrap_cols and col_idx in wrap_cols:
            cell.alignment = _WRAP_ALIGN
        else:
            cell.alignment = _TOP_ALIGN


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def write_excel_report(
    graph: DependencyGraph,
    output_path: Union[str, Path, None] = None,
    exclude_system: bool = False,
) -> bytes:
    """Build and save an Excel workbook from a fully-linked DependencyGraph.

    The workbook is always saved to the user's Downloads folder as
    'dependency_report.xlsx'.  An additional copy is also written to
    ``output_path`` when provided.

    Args:
        graph:          The completed DependencyGraph returned by DependencyEngine.build().
        output_path:    Optional extra save location (e.g. ./output/dependency_report.xlsx).
                        The Downloads copy is always written regardless of this argument.
        exclude_system: When True, skip LocalDateTable_* / DateTableTemplate_* tables.

    Returns:
        The raw bytes of the .xlsx workbook.
    """
    wb = Workbook()
    wb.remove(wb.active)   # remove default empty sheet

    _build_sheet1(wb, graph, exclude_system)
    _build_sheet2(wb, graph, exclude_system)
    _build_sheet3(wb, graph, exclude_system)

    buf = io.BytesIO()
    wb.save(buf)
    raw = buf.getvalue()

    # --- Always save a new timestamped file to the user's Downloads folder ---
    downloads_dir = Path.home() / "Downloads"
    downloads_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    downloads_path = downloads_dir / f"dependency_report_{timestamp}.xlsx"
    downloads_path.write_bytes(raw)
    logger.info("Saved Excel report to Downloads: %s", downloads_path)

    # --- Optional extra copy ---
    if output_path is not None:
        path = Path(output_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_bytes(raw)
        logger.info("Wrote Excel report: %s", path)

    return raw


# ---------------------------------------------------------------------------
# Sheet 1 — Visual Inventory
# ---------------------------------------------------------------------------

def _build_sheet1(wb: Workbook, graph: DependencyGraph, exclude_system: bool) -> None:
    ws = wb.create_sheet("Visual Inventory")

    headers = [
        "Page Name",
        "Visual Name",
        "Visual Type",
        "Direct Measures",
        "Dependency Tables",
        "Dependency Columns",
        "Full DAX of Direct Measures",
    ]
    ws.append(headers)
    _style_header(ws, len(headers))

    # Wrap-text columns: Direct Measures(4), Dep Tables(5), Dep Columns(6), DAX(7)
    WRAP = {4, 5, 6, 7}

    for visual in sorted(graph.visuals.values(), key=lambda v: (v.page, v.title)):
        if exclude_system:
            # Skip visuals that only reference system tables
            non_system = {t for t in visual.tables if not _is_system(t)}
            if not non_system and visual.tables:
                continue

        # Compute dependency tables and columns via trace_measure_lineage
        dep_tables: set[str] = set()
        dep_columns: set[str] = set()

        for measure_name in visual.measures:
            measure = graph.measures.get(measure_name)
            if measure is not None:
                m_tables, m_columns = trace_measure_lineage(measure, graph)
                dep_tables |= m_tables
                dep_columns |= m_columns

        # Add tables/columns used directly by the visual (not through a measure)
        for col_ref in visual.columns:
            dep_columns.add(col_ref)
            table_part = col_ref.partition("[")[0]
            dep_tables.add(table_part)

        if exclude_system:
            dep_tables = {t for t in dep_tables if not _is_system(t)}
            dep_columns = {c for c in dep_columns if not _is_system(c.partition("[")[0])}

        # Build DAX block: label each formula with its measure name
        dax_parts: list[str] = []
        for measure_name in sorted(visual.measures):
            measure = graph.measures.get(measure_name)
            if measure is not None and measure.dax:
                dax_parts.append(f"{measure_name}:\n{measure.dax}")
        dax_text = "\n\n".join(dax_parts)

        _write_row(
            ws,
            ws.max_row + 1,
            [
                visual.page,
                visual.title,
                visual.type,
                _join(visual.measures) or None,
                _join(dep_tables) or None,
                _join(dep_columns) or None,
                dax_text or None,
            ],
            wrap_cols=WRAP,
        )

    _auto_width(ws)


# ---------------------------------------------------------------------------
# Sheet 2 — Impact Analysis
# ---------------------------------------------------------------------------

# Status sort priority — lower value surfaces first
_STATUS_ORDER = {
    "No References Found": 0,
    "Bridge Table": 1,
    "Used via Relationship Only": 2,
    "System Table (auto-generated)": 3,
    "Active Table": 4,
    "Active Measure": 4,
    "Active Column": 4,
}

# Object type sort priority
_TYPE_ORDER = {"Table": 0, "Measure": 1, "Column": 2}


def _build_sheet2(wb: Workbook, graph: DependencyGraph, exclude_system: bool) -> None:
    ws = wb.create_sheet("Impact Analysis")

    headers = [
        "Object Type",
        "Object Name",
        "Home Table",
        "# Visuals Using It",
        "# Measures Depending On It",
        "Used In Relationships (Y/N)",
        "Status",
        "Reason",
    ]
    ws.append(headers)
    _style_header(ws, len(headers))

    WRAP = {8}  # Wrap the Reason column

    # --- Pre-compute lookup structures ---
    unused = find_unused_entities(graph)
    unused_table_set: set[str] = set(unused["unused_tables"])
    unused_measure_set: set[str] = set(unused["unused_measures"])
    unused_column_map: dict[str, list[str]] = unused["unused_columns"]

    # Relationship lookup sets
    rel_tables: set[str] = set()
    rel_columns: set[str] = set()
    for rel in graph.relationships:
        rel_tables.add(rel.from_table)
        rel_tables.add(rel.to_table)
        rel_columns.add(f"{rel.from_table}[{rel.from_column}]")
        rel_columns.add(f"{rel.to_table}[{rel.to_column}]")

    # Count relationships per table (for Bridge Table detection)
    rel_partner_count: dict[str, set[str]] = {}
    for rel in graph.relationships:
        rel_partner_count.setdefault(rel.from_table, set()).add(rel.to_table)
        rel_partner_count.setdefault(rel.to_table, set()).add(rel.from_table)

    # Pre-compute which columns are referenced by which visuals (direct)
    col_to_visual_direct: dict[str, set[str]] = {}
    for visual in graph.visuals.values():
        for col_ref in visual.columns:
            col_to_visual_direct.setdefault(col_ref, set()).add(visual.title)

    # Pre-compute measure lineage for all measures (for column visual/measure counts)
    measure_lineage_cache: dict[str, tuple[set[str], set[str]]] = {}
    for measure_name, measure in graph.measures.items():
        measure_lineage_cache[measure_name] = trace_measure_lineage(measure, graph)

    # For each column: count visuals that reach it through a measure's lineage
    col_to_visual_via_measure: dict[str, set[str]] = {}
    for measure_name, measure in graph.measures.items():
        _, lineage_cols = measure_lineage_cache[measure_name]
        # Strip " (calc)" suffix for matching purposes
        lineage_col_names = set()
        for c in lineage_cols:
            lineage_col_names.add(re.sub(r" \(calc\)$", "", c))
        for col in lineage_col_names:
            for visual_title in measure.visuals:
                col_to_visual_via_measure.setdefault(col, set()).add(visual_title)

    # For each column: count measures whose lineage includes it
    col_to_measures: dict[str, set[str]] = {}
    for measure_name, measure in graph.measures.items():
        _, lineage_cols = measure_lineage_cache[measure_name]
        lineage_col_names = set()
        for c in lineage_cols:
            lineage_col_names.add(re.sub(r" \(calc\)$", "", c))
        for col in lineage_col_names:
            col_to_measures.setdefault(col, set()).add(measure_name)

    # For tables: check if any column appears directly on a visual
    def _table_has_direct_visual_columns(table_name: str) -> bool:
        for col_name in graph.tables[table_name].columns:
            qualified = f"{table_name}[{col_name}]"
            if qualified in col_to_visual_direct:
                return True
        return False

    # --- Collect all rows ---
    rows: list[tuple] = []

    # --- Table rows ---
    for table_name, table in graph.tables.items():
        if exclude_system and _is_system(table_name):
            continue

        n_visuals = len(table.visuals)
        n_measures = len(table.measures)
        in_rel = "Y" if table_name in rel_tables else "N"

        # Status classification
        if _is_system(table_name):
            status = "System Table (auto-generated)"
            reason = "Auto-generated Power BI date/time intelligence table"
        elif (
            len(rel_partner_count.get(table_name, set())) >= 2
            and n_measures == 0
            and not _table_has_direct_visual_columns(table_name)
            and n_visuals == 0
        ):
            status = "Bridge Table"
            reason = f"Joins {len(rel_partner_count[table_name])} tables via relationships, no direct measure or visual references"
        elif table_name in unused_table_set:
            status = "No References Found"
            reason = "No direct or transitive references found"
        else:
            status = "Active Table"
            parts = []
            if n_visuals:
                parts.append(f"Used in {n_visuals} visual{'s' if n_visuals != 1 else ''}")
            if n_measures:
                parts.append(f"{n_measures} measure{'s' if n_measures != 1 else ''}")
            reason = ", ".join(parts) if parts else "Referenced in model"

        rows.append((
            "Table", table_name, "", n_visuals, n_measures, in_rel, status, reason,
        ))

    # --- Measure rows ---
    for measure_name, measure in graph.measures.items():
        if exclude_system and _is_system(measure.table):
            continue

        n_visuals = len(measure.visuals)
        n_used_by = len(measure.used_by_measures)

        if _is_system(measure.table):
            status = "System Table (auto-generated)"
            reason = "Measure on auto-generated date/time table"
        elif measure_name in unused_measure_set:
            status = "No References Found"
            reason = "No direct or transitive references found"
        else:
            status = "Active Measure"
            parts = []
            if n_visuals:
                parts.append(f"Used in {n_visuals} visual{'s' if n_visuals != 1 else ''}")
            if n_used_by:
                parts.append(f"referenced by {n_used_by} other measure{'s' if n_used_by != 1 else ''}")
            reason = ", ".join(parts) if parts else "Referenced in model"

        rows.append((
            "Measure", measure_name, measure.table, n_visuals, n_used_by, "N", status, reason,
        ))

    # --- Column rows ---
    for table_name, table in graph.tables.items():
        if exclude_system and _is_system(table_name):
            continue

        unused_cols_for_table = set(unused_column_map.get(table_name, []))

        for col_name in sorted(table.columns):
            qualified = f"{table_name}[{col_name}]"

            # Count visuals: direct + via measure lineage
            direct_vis = col_to_visual_direct.get(qualified, set())
            measure_vis = col_to_visual_via_measure.get(qualified, set())
            all_vis = direct_vis | measure_vis
            n_visuals = len(all_vis)

            # Count measures whose lineage includes this column
            dep_measures = col_to_measures.get(qualified, set())
            n_measures = len(dep_measures)

            in_rel = "Y" if qualified in rel_columns else "N"

            # Status classification
            if _is_system(table_name):
                status = "System Table (auto-generated)"
                reason = "Column on auto-generated date/time table"
            elif n_visuals == 0 and n_measures == 0 and in_rel == "Y":
                status = "Used via Relationship Only"
                reason = "Only referenced as a relationship join key"
            elif col_name in unused_cols_for_table and in_rel == "N":
                status = "No References Found"
                reason = "No direct or transitive references found"
            elif n_visuals > 0 or n_measures > 0:
                status = "Active Column"
                parts = []
                if n_visuals:
                    parts.append(f"Used in {n_visuals} visual{'s' if n_visuals != 1 else ''}")
                if n_measures:
                    parts.append(f"{n_measures} measure{'s' if n_measures != 1 else ''}")
                reason = ", ".join(parts)
            else:
                # Column not unused per find_unused_entities (e.g. relationship)
                # but also not directly in visuals/measures — relationship key
                if in_rel == "Y":
                    status = "Used via Relationship Only"
                    reason = "Only referenced as a relationship join key"
                else:
                    status = "No References Found"
                    reason = "No direct or transitive references found"

            rows.append((
                "Column", qualified, table_name, n_visuals, n_measures, in_rel, status, reason,
            ))

    # --- Sort: Status priority, then Object Type, then Object Name ---
    rows.sort(key=lambda r: (
        _STATUS_ORDER.get(r[6], 99),
        _TYPE_ORDER.get(r[0], 99),
        r[1],
    ))

    # --- Write rows and apply conditional format ---
    for row_data in rows:
        row_idx = ws.max_row + 1
        _write_row(ws, row_idx, list(row_data), wrap_cols=WRAP)

        # Apply pale highlight to "No References Found" rows
        if row_data[6] == "No References Found":
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = _NO_REF_FILL

    _auto_width(ws)


# ---------------------------------------------------------------------------
# Sheet 3 — Measure Lineage
# ---------------------------------------------------------------------------

def _build_sheet3(wb: Workbook, graph: DependencyGraph, exclude_system: bool) -> None:
    ws = wb.create_sheet("Measure Lineage")

    headers = [
        "Measure Name",
        "Direct Measure Dependencies",
        "Base Columns Used",
        "Dependency Tables",
        "Used in Visuals",
        "Used by Other Measures",
        "Full DAX Expression",
    ]
    ws.append(headers)
    _style_header(ws, len(headers))

    # Wrap all columns except Measure Name (1)
    WRAP = {2, 3, 4, 5, 6, 7}

    for name, measure in sorted(graph.measures.items()):
        if exclude_system and _is_system(measure.table):
            continue

        lineage_tables, lineage_columns = trace_measure_lineage(measure, graph)

        if exclude_system:
            lineage_tables = {t for t in lineage_tables if not _is_system(t)}
            lineage_columns = {
                c for c in lineage_columns
                if not _is_system(re.sub(r" \(calc\)$", "", c).partition("[")[0])
            }

        _write_row(
            ws,
            ws.max_row + 1,
            [
                name,
                _join(measure.depends_on_measures) or None,
                _join(lineage_columns) or None,
                _join(lineage_tables) or None,
                _join(measure.visuals) or None,
                _join(measure.used_by_measures) or None,
                measure.dax or None,
            ],
            wrap_cols=WRAP,
        )

    _auto_width(ws)
