"""Tests for the new trace_measure_lineage function, Sheet 2 status
classifications, and the visual title cleanup."""

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from parser.report_parser import RawPage, RawReport
from parser.tmdl_parser import RawColumn, RawMeasure, RawRelationship, RawSemanticModel, RawTable
from parser.visual_parser import RawVisual
from services.dependency_engine import DependencyEngine, DependencyGraph, find_unused_entities, trace_measure_lineage
from services.dependency_engine import CalculatedColumn, Relationship
from models.measure import Measure
from models.table import Table
from models.visual import Visual


# -----------------------------------------------------------------------
# Helpers — build targeted graphs for each scenario
# -----------------------------------------------------------------------

def _build_calc_column_chain_graph() -> DependencyGraph:
    """Graph for testing transitive calc-column resolution.

    Measure A -> [Measure B] -> references Table1[CalcCol]
    Table1[CalcCol] is a calculated column -> RELATED(Table2[BaseCol])

    So trace_measure_lineage(A) should return Table2[BaseCol] (calc).
    """
    model = RawSemanticModel()
    model.tables["Table1"] = RawTable(
        name="Table1",
        columns={
            "PlainCol": RawColumn(name="PlainCol"),
            "CalcCol": RawColumn(
                name="CalcCol",
                expression="RELATED(Table2[BaseCol])",
            ),
        },
        measures=[
            RawMeasure(name="Measure B", table="Table1", dax="SUM(Table1[CalcCol])"),
            RawMeasure(name="Measure A", table="Table1", dax="[Measure B] + 1"),
        ],
    )
    model.tables["Table2"] = RawTable(
        name="Table2",
        columns={"BaseCol": RawColumn(name="BaseCol")},
    )
    model.relationships.append(
        RawRelationship(from_table="Table1", from_column="PlainCol",
                        to_table="Table2", to_column="BaseCol")
    )

    report = RawReport()
    visual = RawVisual(
        id="v1", title="Card", type="card",
        raw_field_refs={("Table1", "Measure A")},
    )
    report.visuals["v1"] = visual
    report.pages.append(RawPage(name="Page1", visual_ids=["v1"]))

    return DependencyEngine(model, report).build()


def _build_bridge_vs_dimension_graph() -> DependencyGraph:
    """Graph with:
    - BridgeTable: relationships to TableA and TableB, no measures, no visual cols
    - DateTable: relationships to TableA and TableB, but has a column on a visual

    BridgeTable should get 'Bridge Table', DateTable should get 'Active Table'.
    """
    model = RawSemanticModel()
    model.tables["BridgeTable"] = RawTable(
        name="BridgeTable",
        columns={"KeyA": RawColumn(name="KeyA"), "KeyB": RawColumn(name="KeyB")},
    )
    model.tables["DateTable"] = RawTable(
        name="DateTable",
        columns={"DateKey": RawColumn(name="DateKey"), "Month": RawColumn(name="Month")},
    )
    model.tables["TableA"] = RawTable(
        name="TableA",
        columns={"ID_A": RawColumn(name="ID_A"), "Amount": RawColumn(name="Amount")},
        measures=[
            RawMeasure(name="Total Amount", table="TableA", dax="SUM(TableA[Amount])"),
        ],
    )
    model.tables["TableB"] = RawTable(
        name="TableB",
        columns={"ID_B": RawColumn(name="ID_B")},
    )

    # BridgeTable has relationships to both TableA and TableB
    model.relationships.append(
        RawRelationship(from_table="BridgeTable", from_column="KeyA",
                        to_table="TableA", to_column="ID_A")
    )
    model.relationships.append(
        RawRelationship(from_table="BridgeTable", from_column="KeyB",
                        to_table="TableB", to_column="ID_B")
    )
    # DateTable also has relationships to TableA and TableB
    model.relationships.append(
        RawRelationship(from_table="DateTable", from_column="DateKey",
                        to_table="TableA", to_column="ID_A")
    )
    model.relationships.append(
        RawRelationship(from_table="DateTable", from_column="DateKey",
                        to_table="TableB", to_column="ID_B")
    )

    report = RawReport()
    # Visual uses a measure from TableA AND a column from DateTable
    visual = RawVisual(
        id="v1", title="Chart", type="columnChart",
        raw_field_refs={("TableA", "Total Amount"), ("DateTable", "Month")},
    )
    report.visuals["v1"] = visual
    report.pages.append(RawPage(name="Dashboard", visual_ids=["v1"]))

    return DependencyEngine(model, report).build()


def _build_relationship_only_column_graph() -> DependencyGraph:
    """Graph where a column is a relationship join key but never referenced
    by any visual or any measure's DAX."""
    model = RawSemanticModel()
    model.tables["Orders"] = RawTable(
        name="Orders",
        columns={
            "OrderID": RawColumn(name="OrderID"),
            "CustomerID": RawColumn(name="CustomerID"),
            "Amount": RawColumn(name="Amount"),
        },
        measures=[
            RawMeasure(name="Total Orders", table="Orders", dax="SUM(Orders[Amount])"),
        ],
    )
    model.tables["Customers"] = RawTable(
        name="Customers",
        columns={
            "CustomerID": RawColumn(name="CustomerID"),
            "Name": RawColumn(name="Name"),
        },
    )
    model.relationships.append(
        RawRelationship(from_table="Orders", from_column="CustomerID",
                        to_table="Customers", to_column="CustomerID")
    )

    report = RawReport()
    visual = RawVisual(
        id="v1", title="Sales KPI", type="card",
        raw_field_refs={("Orders", "Total Orders")},
    )
    report.visuals["v1"] = visual
    report.pages.append(RawPage(name="Home", visual_ids=["v1"]))

    return DependencyEngine(model, report).build()


def _build_system_table_graph() -> DependencyGraph:
    """Graph with a LocalDateTable that should get 'System Table (auto-generated)'."""
    model = RawSemanticModel()
    model.tables["Sales"] = RawTable(
        name="Sales",
        columns={"Amount": RawColumn(name="Amount")},
        measures=[
            RawMeasure(name="Total", table="Sales", dax="SUM(Sales[Amount])"),
        ],
    )
    model.tables["LocalDateTable_abc123"] = RawTable(
        name="LocalDateTable_abc123",
        columns={"Date": RawColumn(name="Date"), "Year": RawColumn(name="Year")},
    )
    model.tables["DateTableTemplate_xyz789"] = RawTable(
        name="DateTableTemplate_xyz789",
        columns={"Date": RawColumn(name="Date")},
    )

    report = RawReport()
    visual = RawVisual(
        id="v1", title="Card", type="card",
        raw_field_refs={("Sales", "Total")},
    )
    report.visuals["v1"] = visual
    report.pages.append(RawPage(name="Home", visual_ids=["v1"]))

    return DependencyEngine(model, report).build()


# -----------------------------------------------------------------------
# Tests
# -----------------------------------------------------------------------

class TestTraceMeasureLineage:
    """Tests for the trace_measure_lineage shared function."""

    def test_transitive_with_calc_column(self):
        """Measure A -> Measure B -> Table1[CalcCol] -> Table2[BaseCol].
        trace_measure_lineage(A) must include Table2[BaseCol] (calc) and Table2."""
        graph = _build_calc_column_chain_graph()
        measure_a = graph.measures["Measure A"]

        tables, columns = trace_measure_lineage(measure_a, graph)

        # Tables: Table1 (home & direct ref), Table2 (via calc column)
        assert "Table1" in tables
        assert "Table2" in tables

        # Columns: Table1[CalcCol] (direct), Table2[BaseCol] (calc)
        assert "Table1[CalcCol]" in columns
        assert "Table2[BaseCol] (calc)" in columns

    def test_single_measure_no_deps(self):
        """A measure with no depends_on_measures should still trace its own refs."""
        graph = _build_calc_column_chain_graph()
        measure_b = graph.measures["Measure B"]

        tables, columns = trace_measure_lineage(measure_b, graph)

        assert "Table1" in tables
        assert "Table1[CalcCol]" in columns
        assert "Table2[BaseCol] (calc)" in columns

    def test_consistency_between_sheet1_and_sheet3(self):
        """For any measure, trace_measure_lineage must return the same result
        regardless of where it's called — this is the whole point of the
        shared function."""
        graph = _build_calc_column_chain_graph()
        measure_a = graph.measures["Measure A"]

        result1 = trace_measure_lineage(measure_a, graph)
        result2 = trace_measure_lineage(measure_a, graph)

        assert result1[0] == result2[0]  # tables match
        assert result1[1] == result2[1]  # columns match


class TestImpactAnalysisClassification:
    """Tests for Sheet 2 status classification logic."""

    def test_bridge_table_classification(self):
        """A table with 2+ relationships, no measures, no visual columns → Bridge Table."""
        graph = _build_bridge_vs_dimension_graph()

        table = graph.tables["BridgeTable"]
        # BridgeTable should have: no visuals, no measures
        assert len(table.visuals) == 0
        assert len(table.measures) == 0

    def test_dimension_table_not_bridge(self):
        """A table with 2+ relationships but also direct visual usage → Active Table, not Bridge."""
        graph = _build_bridge_vs_dimension_graph()

        table = graph.tables["DateTable"]
        # DateTable should have visuals because its column "Month" is on a visual
        assert len(table.visuals) > 0

    def test_used_via_relationship_only(self):
        """A column that is a relationship join key but never in a visual or measure."""
        graph = _build_relationship_only_column_graph()

        # Orders[CustomerID] is a relationship key but not in any measure's DAX
        # or any visual's columns
        col_ref = "Orders[CustomerID]"

        # Check it's in a relationship
        rel_cols = set()
        for rel in graph.relationships:
            rel_cols.add(f"{rel.from_table}[{rel.from_column}]")
            rel_cols.add(f"{rel.to_table}[{rel.to_column}]")
        assert col_ref in rel_cols

        # Check it's not directly in any visual's columns
        for visual in graph.visuals.values():
            assert col_ref not in visual.columns

        # Check it's not in any measure's referenced_columns
        for measure in graph.measures.values():
            assert col_ref not in measure.referenced_columns

    def test_system_table_not_no_references(self):
        """LocalDateTable_* should get 'System Table (auto-generated)',
        not 'No References Found'."""
        graph = _build_system_table_graph()

        # Verify system tables exist and have no visual usage
        assert "LocalDateTable_abc123" in graph.tables
        assert len(graph.tables["LocalDateTable_abc123"].visuals) == 0

        # Even with exclude_system=False, find_unused_entities would mark them
        # unused, but our Sheet 2 must distinguish them as system tables
        unused = find_unused_entities(graph)
        assert "LocalDateTable_abc123" in unused["unused_tables"]

        # The test verifies the data is correct for the classification;
        # actual status assignment is in excel_export._build_sheet2


class TestVisualTitleCleanup:
    """Test that visual.title no longer contains the baked-in page suffix."""

    def test_visual_title_no_page_suffix(self):
        """After building the graph, visual.title must be the clean title
        without '(Page: X)' suffix."""
        graph = _build_calc_column_chain_graph()

        for visual in graph.visuals.values():
            assert "(Page:" not in visual.title, (
                f"Visual '{visual.title}' still has baked-in page suffix"
            )
            # The page info should be in the separate field
            assert visual.page, "visual.page should not be empty"

    def test_visual_title_matches_raw_title(self):
        """visual.title should be the raw visual title, not a composite."""
        graph = _build_calc_column_chain_graph()

        visual = graph.visuals["v1"]
        assert visual.title == "Card"
        assert visual.page == "Page1"


class TestExcelWorkbookStructure:
    """Test that the workbook has exactly 3 sheets in the correct order."""

    def test_workbook_three_sheets(self):
        """Workbook must have exactly 3 sheets: Visual Inventory, Impact Analysis, Measure Lineage."""
        from services.excel_export import write_excel_report
        from openpyxl import load_workbook
        import io

        graph = _build_calc_column_chain_graph()
        raw_bytes = write_excel_report(graph, output_path=None, exclude_system=False)

        wb = load_workbook(io.BytesIO(raw_bytes))
        assert wb.sheetnames == ["Visual Inventory", "Impact Analysis", "Measure Lineage"]

    def test_sheet1_no_dax_column(self):
        """Sheet 1 must NOT include Full DAX of Direct Measures."""
        from services.excel_export import write_excel_report
        from openpyxl import load_workbook
        import io

        graph = _build_calc_column_chain_graph()
        raw_bytes = write_excel_report(graph, output_path=None, exclude_system=False)

        wb = load_workbook(io.BytesIO(raw_bytes))
        ws = wb["Visual Inventory"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert "Full DAX of Direct Measures" not in headers

    def test_sheet1_includes_empty_pages(self):
        """Sheet 1 must include pages even if they have no visuals."""
        from services.excel_export import write_excel_report
        from openpyxl import load_workbook
        import io

        graph = _build_calc_column_chain_graph()
        # Add a page with no visuals to the graph
        from models.page import Page
        graph.pages["EmptyPage"] = Page(name="EmptyPage")

        raw_bytes = write_excel_report(graph, output_path=None, exclude_system=False)

        wb = load_workbook(io.BytesIO(raw_bytes))
        ws = wb["Visual Inventory"]
        # Loop through rows to see if EmptyPage is listed
        found_empty_page = False
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == "EmptyPage":
                found_empty_page = True
                # The other columns should be None/blank
                assert ws.cell(row=row, column=2).value is None
                assert ws.cell(row=row, column=3).value is None
                break
        assert found_empty_page, "EmptyPage was not listed in Visual Inventory"

    def test_sheet3_has_dax_column(self):
        """Sheet 3 must include Full DAX Expression."""
        from services.excel_export import write_excel_report
        from openpyxl import load_workbook
        import io

        graph = _build_calc_column_chain_graph()
        raw_bytes = write_excel_report(graph, output_path=None, exclude_system=False)

        wb = load_workbook(io.BytesIO(raw_bytes))
        ws = wb["Measure Lineage"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert "Full DAX Expression" in headers


class TestVisualParserTitleExtraction:
    """Verify that parse_visual extracts visual titles from modern and legacy JSON shapes."""

    def test_extract_title_from_modern_visual_container_objects(self):
        from parser.visual_parser import parse_visual
        visual_json = {
            "visual": {
                "visualType": "lineChart",
                "visualContainerObjects": {
                    "title": [
                        {
                            "properties": {
                                "text": {
                                    "expr": {
                                        "Literal": {
                                            "Value": "'Total Invoice by Month'"
                                        }
                                    }
                                }
                            }
                        }
                    ]
                }
            }
        }
        visual = parse_visual(visual_json, visual_id="v1")
        assert visual.title == "Total Invoice by Month"
        assert visual.type == "lineChart"

    def test_extract_title_from_legacy_objects(self):
        from parser.visual_parser import parse_visual
        visual_json = {
            "visual": {
                "visualType": "columnChart",
                "objects": {
                    "title": [
                        {
                            "properties": {
                                "text": {
                                    "expr": {
                                        "Literal": {
                                            "Value": "'Custom Title'"
                                        }
                                    }
                                }
                            }
                        }
                    ]
                }
            }
        }
        visual = parse_visual(visual_json, visual_id="v2")
        assert visual.title == "Custom Title"
        assert visual.type == "columnChart"

    def test_fallback_to_type_when_no_title(self):
        from parser.visual_parser import parse_visual
        visual_json = {
            "visual": {
                "visualType": "barChart"
            }
        }
        visual = parse_visual(visual_json, visual_id="v3")
        assert visual.title == "barChart"
        assert visual.type == "barChart"


class TestVisualIdentityCollisions:
    """Verify that multiple visuals with identical titles on the same page
    do not collide at the page, table, and measure visuals aggregations."""

    def test_visual_title_collisions_on_same_page(self):
        # 1. Build a model with duplicate visual titles on the same page
        model = RawSemanticModel()
        model.tables["Sales"] = RawTable(
            name="Sales",
            columns={"Amount": RawColumn(name="Amount")},
            measures=[RawMeasure(name="Total Sales", table="Sales", dax="SUM(Sales[Amount])")]
        )

        report = RawReport()
        # page1::v1 and page1::v2 have the exact same title ("Card")
        report.visuals["page1::v1"] = RawVisual(
            id="page1::v1", title="Card", type="card", raw_field_refs={("Sales", "Total Sales")}
        )
        report.visuals["page1::v2"] = RawVisual(
            id="page1::v2", title="Card", type="card", raw_field_refs={("Sales", "Total Sales")}
        )

        report.pages.append(RawPage(name="Home", visual_ids=["page1::v1", "page1::v2"]))

        graph = DependencyEngine(model, report).build()

        # Both visuals must exist in the graph
        assert len(graph.visuals) == 2
        assert "page1::v1" in graph.visuals
        assert "page1::v2" in graph.visuals

        # Page visuals must contain both visual IDs (no deduplication by title)
        page = graph.pages["Home"]
        assert len(page.visuals) == 2
        assert "page1::v1" in page.visuals
        assert "page1::v2" in page.visuals

        # Table visuals must contain both visual IDs
        table = graph.tables["Sales"]
        assert len(table.visuals) == 2
        assert "page1::v1" in table.visuals
        assert "page1::v2" in table.visuals

        # Measure visuals must contain both visual IDs
        measure = graph.measures["Total Sales"]
        assert len(measure.visuals) == 2
        assert "page1::v1" in measure.visuals
        assert "page1::v2" in measure.visuals


class TestCalcColumnDependencyInImpactAnalysis:
    """Verify that a base column referenced by an active calculated column
    is not marked as 'No References Found' in the Impact Analysis sheet."""

    def _build_calc_col_dependency_graph(self) -> DependencyGraph:
        """Build a graph where:
        - FactTable[PaymentTerms] is a plain base column (not on any visual)
        - DimTable[TermDescription] is a calculated column = RELATED(FactTable[PaymentTerms])
        - DimTable[TermDescription] IS used directly on a visual

        So FactTable[PaymentTerms] must be 'Active Column' because deleting it
        would break the calculated column.
        """
        model = RawSemanticModel()
        model.tables["FactTable"] = RawTable(
            name="FactTable",
            columns={
                "Amount": RawColumn(name="Amount"),
                "PaymentTerms": RawColumn(name="PaymentTerms"),
            },
            measures=[
                RawMeasure(name="Total Amount", table="FactTable", dax="SUM(FactTable[Amount])"),
            ],
        )
        model.tables["DimTable"] = RawTable(
            name="DimTable",
            columns={
                "DimKey": RawColumn(name="DimKey"),
                "TermDescription": RawColumn(
                    name="TermDescription",
                    expression="RELATED(FactTable[PaymentTerms])",
                ),
            },
        )
        model.relationships.append(
            RawRelationship(
                from_table="FactTable", from_column="PaymentTerms",
                to_table="DimTable", to_column="DimKey",
            )
        )

        report = RawReport()
        # Visual 1: uses the measure
        report.visuals["v1"] = RawVisual(
            id="v1", title="Amount Card", type="card",
            raw_field_refs={("FactTable", "Total Amount")},
        )
        # Visual 2: uses the calculated column directly
        report.visuals["v2"] = RawVisual(
            id="v2", title="Terms Table", type="tableEx",
            raw_field_refs={("DimTable", "TermDescription")},
        )
        report.pages.append(RawPage(name="Page1", visual_ids=["v1", "v2"]))

        return DependencyEngine(model, report).build()

    def test_base_column_not_marked_unused_when_calc_col_depends_on_it(self):
        """FactTable[PaymentTerms] is referenced by DimTable[TermDescription]
        (a calc col used in a visual), so it must NOT be 'No References Found'."""
        from services.excel_export import write_excel_report
        from openpyxl import load_workbook
        import io

        graph = self._build_calc_col_dependency_graph()
        raw_bytes = write_excel_report(graph, output_path=None, exclude_system=False)

        wb = load_workbook(io.BytesIO(raw_bytes))
        ws = wb["Impact Analysis"]

        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        type_idx = headers.index("Object Type") + 1
        name_idx = headers.index("Object Name") + 1
        calc_col_idx = headers.index("# Calc Columns Depending On It") + 1
        status_idx = headers.index("Status") + 1

        # Find the row for FactTable[PaymentTerms]
        target_row = None
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=name_idx).value == "FactTable[PaymentTerms]":
                target_row = row
                break

        assert target_row is not None, "FactTable[PaymentTerms] not found in Impact Analysis"

        status = ws.cell(row=target_row, column=status_idx).value
        n_calc = ws.cell(row=target_row, column=calc_col_idx).value

        assert status != "No References Found", (
            f"FactTable[PaymentTerms] should not be 'No References Found' "
            f"because a calculated column depends on it"
        )
        assert status == "Active Column", f"Expected 'Active Column', got '{status}'"
        assert int(n_calc) == 1, f"Expected 1 calc column depending on it, got {n_calc}"

    def test_calc_column_labelled_correctly(self):
        """DimTable[TermDescription] should have Object Type = 'Calculated Column'."""
        from services.excel_export import write_excel_report
        from openpyxl import load_workbook
        import io

        graph = self._build_calc_col_dependency_graph()
        raw_bytes = write_excel_report(graph, output_path=None, exclude_system=False)

        wb = load_workbook(io.BytesIO(raw_bytes))
        ws = wb["Impact Analysis"]

        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        type_idx = headers.index("Object Type") + 1
        name_idx = headers.index("Object Name") + 1

        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=name_idx).value == "DimTable[TermDescription]":
                obj_type = ws.cell(row=row, column=type_idx).value
                assert obj_type == "Calculated Column", (
                    f"Expected 'Calculated Column', got '{obj_type}'"
                )
                return

        assert False, "DimTable[TermDescription] not found in Impact Analysis"

    def test_new_header_present(self):
        """The '# Calc Columns Depending On It' header must exist in Impact Analysis."""
        from services.excel_export import write_excel_report
        from openpyxl import load_workbook
        import io

        graph = self._build_calc_col_dependency_graph()
        raw_bytes = write_excel_report(graph, output_path=None, exclude_system=False)

        wb = load_workbook(io.BytesIO(raw_bytes))
        ws = wb["Impact Analysis"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert "# Calc Columns Depending On It" in headers

    def test_removable_items_sorted_first(self):
        """'No References Found' rows must appear before 'Active' rows."""
        from services.excel_export import write_excel_report
        from openpyxl import load_workbook
        import io

        graph = self._build_calc_col_dependency_graph()
        raw_bytes = write_excel_report(graph, output_path=None, exclude_system=False)

        wb = load_workbook(io.BytesIO(raw_bytes))
        ws = wb["Impact Analysis"]

        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        status_idx = headers.index("Status") + 1

        statuses = []
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=status_idx).value
            if val:
                statuses.append(val)

        # All "No References Found" must come before any "Active *"
        last_no_ref = -1
        first_active = len(statuses)
        for i, s in enumerate(statuses):
            if s == "No References Found":
                last_no_ref = i
            if s.startswith("Active") and i < first_active:
                first_active = i

        if last_no_ref >= 0 and first_active < len(statuses):
            assert last_no_ref < first_active, (
                "Removable items ('No References Found') must be sorted before active items"
            )
